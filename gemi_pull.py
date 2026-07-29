# -*- coding: utf-8 -*-
"""
ΓΕΜΗ OpenData -> Excel ανά ΚΑΔ (ένα φύλλο/tab ανά Νομό).

Χρήση:
    python gemi_pull.py search ΚΟΜΜΩΤΗΡ    # βρες ΚΑΔ με λέξη-κλειδί
    python gemi_pull.py                     # τρέξε την εξαγωγή (βλ. ρυθμίσεις πιο κάτω)
"""
import os
import sys
import time

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

API_KEY = os.environ.get("GEMI_API_KEY", "fbhvTXTHzr4tpdIvbotQSb8M6eOUiSyY").strip()
BASE = "https://opendata-api.businessportal.gr/api/opendata/v1"
PAGE = 200
SLEEP = 7.8  # ~8 αιτήματα/λεπτό

# ================== ΡΥΘΜΙΣΕΙΣ ΠΟΥ ΑΛΛΑΖΕΙΣ ΕΣΥ ==================
# 1) Ποιους ΚΑΔ θέλεις (βάλε όσους θες — παράγεται 1 xlsx ανά ΚΑΔ):
KADS = ["47730000", "47750000"]  # π.χ. φαρμακεία, καλλυντικά

# 2) Ποιοι Νομοί; "ALL" για όλη την Ελλάδα, ή λίστα με ονόματα/κωδικούς:
PREFECTURES = "ALL"  # π.χ. ["ΘΕΣΣΑΛΟΝΙΚΗΣ", "ΑΤΤΙΚΗΣ"] ή ["19", "5"]

# 3) Κράτα μόνο όσες έχουν τον ΚΑΔ ως ΚΥΡΙΑ δραστηριότητα (καθαρή λίστα):
PRIMARY_ONLY = True

# 4) Αλυσίδες ενός brand που ΔΕΝ θέλεις (case-insensitive, ελέγχει "περιέχει"):
EXCLUDE = [
    "LA VIE EN ROSE", "OLIVE ERA", "DUST+CREAM", "DUST AND CREAM",
    "NYX", "SEPHORA", "HONDOS", "ΧΟΝΔΟΣ", "MAC COSMETICS", "THE BODY SHOP",
    "LUSH", "YVES ROCHER", "KIKO", "INGLOT", "RITUALS", "MASUTTI",
]
# ================================================================

# Πίνακας Νομών (όνομα -> κωδικός) — από το ίδιο το API:
PREF_MAP = {
    "ΑΘΗΝΩΝ": "54", "ΑΙΤΩΛΟΑΚΑΡΝΑΝΙΑΣ": "1", "ΑΝΑΤΟΛΙΚΗΣ ΑΤΤΙΚΗΣ": "52", "ΑΡΓΟΛΙΔΑΣ": "2",
    "ΑΡΚΑΔΙΑΣ": "3", "ΑΡΤΑΣ": "4", "ΑΤΤΙΚΗΣ": "5", "ΑΧΑΙΑΣ": "6", "ΒΟΙΩΤΙΑΣ": "7", "ΓΡΕΒΕΝΩΝ": "8",
    "ΔΡΑΜΑΣ": "9", "ΔΥΤΙΚΗΣ ΑΤΤΙΚΗΣ": "53", "ΔΩΔΕΚΑΝΗΣΟΥ": "10", "ΕΒΡΟΥ": "11", "ΕΥΒΟΙΑΣ": "12",
    "ΕΥΡΥΤΑΝΙΑΣ": "13", "ΖΑΚΥΝΘΟΥ": "14", "ΗΛΕΙΑΣ": "15", "ΗΜΑΘΙΑΣ": "16", "ΗΡΑΚΛΕΙΟΥ": "17",
    "ΘΕΣΠΡΩΤΙΑΣ": "18", "ΘΕΣΣΑΛΟΝΙΚΗΣ": "19", "ΙΩΑΝΝΙΝΩΝ": "20", "ΚΑΒΑΛΑΣ": "21", "ΚΑΡΔΙΤΣΑΣ": "22",
    "ΚΑΣΤΟΡΙΑΣ": "23", "ΚΕΡΚΥΡΑΣ": "24", "ΚΕΦΑΛΛΗΝΙΑΣ": "25", "ΚΙΛΚΙΣ": "26", "ΚΟΖΑΝΗΣ": "27",
    "ΚΟΡΙΝΘΙΑΣ": "28", "ΚΥΚΛΑΔΩΝ": "29", "ΛΑΚΩΝΙΑΣ": "30", "ΛΑΡΙΣΑΣ": "31", "ΛΑΣΙΘΙΟΥ": "32",
    "ΛΕΣΒΟΥ": "33", "ΛΕΥΚΑΔΑΣ": "34", "ΜΑΓΝΗΣΙΑΣ": "35", "ΜΕΣΣΗΝΙΑΣ": "36", "ΞΑΝΘΗΣ": "37",
    "ΠΕΙΡΑΙΑ": "55", "ΠΕΛΛΗΣ": "38", "ΠΙΕΡΙΑΣ": "39", "ΠΡΕΒΕΖΑΣ": "40", "ΡΕΘΥΜΝΗΣ": "41",
    "ΡΟΔΟΠΗΣ": "42", "ΣΑΜΟΥ": "43", "ΣΕΡΡΩΝ": "44", "ΤΡΙΚΑΛΩΝ": "45", "ΦΘΙΩΤΙΔΑΣ": "46",
    "ΦΛΩΡΙΝΑΣ": "47", "ΦΩΚΙΔΑΣ": "48", "ΧΑΛΚΙΔΙΚΗΣ": "49", "ΧΑΝΙΩΝ": "50", "ΧΙΟΥ": "51",
}

HDR = ["Επωνυμία", "ΑΦΜ", "Διεύθυνση", "Πόλη", "ΤΚ", "Νομός", "Τηλέφωνο", "Email", "Website"]
NAVY = "1F3864"


def session():
    s = requests.Session()
    s.headers.update({"api_key": API_KEY, "Accept": "application/json"})
    return s


# ---- ΑΝΑΖΗΤΗΣΗ ΚΑΔ ΜΕ ΛΕΞΗ-ΚΛΕΙΔΙ ----
def search_kad(keyword):
    s = session()
    r = s.get(f"{BASE}/metadata/activities", timeout=60)
    r.raise_for_status()
    kw = keyword.upper()
    hits = [a for a in r.json() if kw in (a.get("descr", "") or "").upper()]
    print(f"Βρέθηκαν {len(hits)} ΚΑΔ για '{keyword}':")
    for a in sorted(hits, key=lambda x: x.get("id", "")):
        print(f"  {a.get('id')}  {a.get('descr')}")


def resolve_prefectures():
    if PREFECTURES == "ALL" or not PREFECTURES:
        return None
    ids = []
    for p in PREFECTURES:
        ids.append(PREF_MAP.get(str(p).upper().strip(), str(p).strip()))
    return ",".join(ids)


def is_primary(company, prefix):
    for a in (company.get("activities") or []):
        act = a.get("activity") or {}
        if a.get("type") == "Κύρια" and str(act.get("id", "")).startswith(prefix):
            return True
    return False


def excluded(name):
    up = (name or "").upper()
    return any(t in up for t in EXCLUDE)


def fetch_kad(kad, pref_ids):
    prefix = kad[:4]
    rows, off, total, seen = [], 0, None, set()
    s = session()
    while total is None or off < total:
        url = f"{BASE}/companies?activities={kad}&isActive=true&resultsSize={PAGE}&resultsOffset={off}"
        if pref_ids:
            url += f"&prefectures={pref_ids}"
        try:
            r = s.get(url, timeout=60)
        except Exception as e:
            print("  net error, retry:", e)
            time.sleep(8)
            continue
        if r.status_code == 429:
            print("  429, wait...")
            time.sleep(9)
            continue
        if r.status_code != 200:
            print("  status", r.status_code, "retry")
            time.sleep(9)
            continue
        j = r.json()
        total = j.get("searchMetadata", {}).get("totalCount", 0)
        arr = j.get("searchResults", []) or []
        for c in arr:
            if PRIMARY_ONLY and not is_primary(c, prefix):
                continue
            name = (c.get("coNameEl") or "").strip()
            if excluded(name):
                continue
            afm = c.get("afm") or ""
            if afm and afm in seen:
                continue
            if afm:
                seen.add(afm)
            pr = c.get("prefecture") or {}
            pr = pr.get("descr", "") if isinstance(pr, dict) else (pr or "")
            addr = ((c.get("street") or "") + " " + (c.get("streetNumber") or "")).strip()
            rows.append([
                name, afm, addr, c.get("city") or "", c.get("zipCode") or "",
                pr, c.get("phone") or "", c.get("email") or "", c.get("url") or "",
            ])
        off += len(arr)
        print(f"  {kad}: kept {len(rows)} / scanned {off}/{total}")
        if len(arr) < PAGE:
            break
        time.sleep(SLEEP)
    return rows


def sanitize(name, used):
    x = (name or "ΛΟΙΠΑ").translate({ord(ch): "-" for ch in ':\\/?*[]'})[:28] or "ΛΟΙΠΑ"
    b, i = x, 1
    while x in used:
        x = f"{b}_{i}"
        i += 1
    used.add(x)
    return x


def build_xlsx(rows, path):
    groups = {}
    for r in rows:
        n = (r[5] or "ΛΟΙΠΑ").strip() or "ΛΟΙΠΑ"
        groups.setdefault(n, []).append(r)
    names = sorted(groups, key=lambda n: -len(groups[n]))
    wb = Workbook()
    ws = wb.active
    ws.title = "Σύνοψη"
    ws.append(["Νομός", "Πλήθος"])
    for n in names:
        ws.append([n, len(groups[n])])
    ws.append(["ΣΥΝΟΛΟ", len(rows)])
    hf = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor=NAVY)
    for c in ws[1]:
        c.font = hf
        c.fill = fill
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 10
    used = set()
    widths = [44, 11, 30, 16, 7, 16, 12, 26, 20]
    for n in names:
        sh = wb.create_sheet(sanitize(n, used))
        sh.append(HDR)
        for c in sh[1]:
            c.font = hf
            c.fill = fill
            c.alignment = Alignment(horizontal="center")
        for row in groups[n]:
            sh.append(row)
        for i, w in enumerate(widths, 1):
            sh.column_dimensions[get_column_letter(i)].width = w
        sh.freeze_panes = "A2"
    wb.save(path)


def main():
    if len(sys.argv) >= 3 and sys.argv[1] == "search":
        search_kad(" ".join(sys.argv[2:]))
        return
    os.makedirs("output", exist_ok=True)
    pref_ids = resolve_prefectures()
    for kad in KADS:
        print("== ΚΑΔ", kad, "| Νομοί:", PREFECTURES)
        rows = fetch_kad(kad, pref_ids)
        out = os.path.join("output", f"GEMH_{kad}_ana_Nomo.xlsx")
        build_xlsx(rows, out)
        print("  -> αποθηκεύτηκε:", out, "με", len(rows), "εγγραφές")


if __name__ == "__main__":
    main()
