# -*- coding: utf-8 -*-
"""
Πυρήνας λογικής για την άντληση δεδομένων από το OpenData API του ΓΕΜΗ
και την παραγωγή Excel (ένα φύλλο/tab ανά Νομό).

Το module αυτό είναι ανεξάρτητο από UI: το χρησιμοποιούν τόσο το CLI
(gemi_pull.py) όσο και η γραφική εφαρμογή (gemi_app.py).
"""
import os
import time
import unicodedata

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = "https://opendata-api.businessportal.gr/api/opendata/v1"
DEFAULT_API_KEY = "fbhvTXTHzr4tpdIvbotQSb8M6eOUiSyY"
PAGE = 200
SLEEP = 7.8  # ~8 αιτήματα/λεπτό

# Προεπιλεγμένη λίστα αποκλεισμού brand/αλυσίδων (case-insensitive, «περιέχει»):
DEFAULT_EXCLUDE = [
    "LA VIE EN ROSE", "OLIVE ERA", "DUST+CREAM", "DUST AND CREAM",
    "NYX", "SEPHORA", "HONDOS", "ΧΟΝΔΟΣ", "MAC COSMETICS", "THE BODY SHOP",
    "LUSH", "YVES ROCHER", "KIKO", "INGLOT", "RITUALS", "MASUTTI",
]

# Πίνακας Νομών (όνομα -> κωδικός) — από το ίδιο το API (55 νομοί):
PREF_MAP = {
    "ΑΘΗΝΩΝ": "54", "ΑΙΤΩΛΟΑΚΑΡΝΑΝΙΑΣ": "1", "ΑΝΑΤΟΛΙΚΗΣ ΑΤΤΙΚΗΣ": "52", "ΑΡΓΟΛΙΔΑΣ": "2",
    "ΑΡΚΑΔΙΑΣ": "3", "ΑΡΤΑΣ": "4", "ΑΤΤΙΚΗΣ": "5", "ΑΧΑΙΑΣ": "6", "ΒΟΙΩΤΙΑΣ": "7", "ΓΡΕΒΕΝΩΝ": "8",
    "ΔΡΑΜΑΣ": "9", "ΔΥΤΙΚΗΣ ΑΤΤΙΚΗΣ": "53", "ΔΩΔΕΚΑΝΗΣΟΥ": "10", "ΕΒΡΟΥ": "11", "ΕΥΒΟΙΑΣ": "12",
    "ΕΥΡΥΤΑΝΙΑΣ": "13", "ΖΑΚΥΝΘΟΥ": "14", "ΗΛΕΙΑΣ": "15", "ΗΜΑΘΙΑΣ": "16", "ΗΡΑΚΛΕΙΟΥ": "17",
    "ΘΕΣΠΡΩΤΙΑΣ": "18", "ΘΕΣΣΑΛΟΝΙΚΗΣ": "19", "ΙΩΑΝΝΙΝΩΝ": "20", "ΚΑΒΑΛΑΣ": "21", "ΚΑΡΔΙΤΣΑΣ": "22",
    "ΚΑΣΤΟΡΙΑΣ": "23", "ΚΕΡΚΥΡΑΣ": "24", "ΚΕΦΑΛΛΗΝΙΑΣ": "25", "ΚΙΛΚΙΣ": "26", "ΚΟΖΑΝΗΣ": "27",
    "ΚΟΡΙΝΘΙΑΣ": "28", "ΚΥΚΛΑΔΩΝ": "29", "ΛΑΚΩΝΙΑΣ": "30", "ΛΑΡΙΣΑΣ": "31", "ΛΑΣΙΘΙΟΥ": "32",
    "ΛΕΣΒΟΥ": "33", "ΛΕΥΚΑΔΑΣ": "34", "ΜΑΓΝΗΣΙΑΣ": "35", "ΜΕΣΣΗΝΙΑΣ": "36", "ΞΑΝΘΗΣ": "37",
    "ΠΕΙΡΑΙΑ": "55", "ΠΕΛΛΗΣ": "38", "ΠΙΕΡΙΑΣ": "39", "ΠΡΕΒΕΖΑΣ": "40", "ΡΕΘΥΜΝΗΣ": "41",
    "ΡΟΔΟΠΗΣ": "42", "ΣΑΜΟΥ": "43", "ΣΕΡΡΩΝ": "44", "ΤΡΙΚΑΛΩΝ": "45", "ΦΘΙΩΤΙΔΑΣ": "46",
    "ΦΛΩΡΙΝΑΣ": "47", "ΦΩΚΙΔΑΣ": "48", "ΧΑΛΚΙΔΙΚΗΣ": "49", "ΧΑΝΙΩΝ": "50", "ΧΙΟΥ": "51",
}
PREF_NAMES = sorted(PREF_MAP.keys())

HDR = ["Επωνυμία", "ΑΦΜ", "Διεύθυνση", "Πόλη", "ΤΚ", "Νομός", "Τηλέφωνο", "Email", "Website"]
NAVY = "1F3864"

# Τρέχουσα ταξινομία ΚΑΔ (το API επιστρέφει και παλιές εκδόσεις).
KAD_VERSION = "kad_2026"


def strip_acc(s):
    """Πεζά/κεφαλαία χωρίς τόνους — για ανεκτική αναζήτηση (π.χ. ΚΟΜΜΩΤΗΡ)."""
    s = unicodedata.normalize("NFD", s or "")
    return "".join(c for c in s if unicodedata.category(c) != "Mn").upper()


# Ελληνικά -> Λατινικά (για ονόματα αρχείων)
_GR2LAT = {
    "Α": "A", "Β": "V", "Γ": "G", "Δ": "D", "Ε": "E", "Ζ": "Z", "Η": "I",
    "Θ": "TH", "Ι": "I", "Κ": "K", "Λ": "L", "Μ": "M", "Ν": "N", "Ξ": "X",
    "Ο": "O", "Π": "P", "Ρ": "R", "Σ": "S", "Τ": "T", "Υ": "Y", "Φ": "F",
    "Χ": "CH", "Ψ": "PS", "Ω": "O", "Ϊ": "I", "Ϋ": "Y",
}


def latin_slug(text, max_len=40, max_words=5):
    """
    Μετατρέπει ελληνικό κείμενο σε λατινικό slug για όνομα αρχείου
    (π.χ. 'ΛΙΑΝΙΚΟ ΕΜΠΟΡΙΟ ΦΑΡΜΑΚΩΝ' -> 'LIANIKO_EMPORIO_FARMAKON').
    """
    up = strip_acc(text)  # κεφαλαία, χωρίς τόνους
    out = []
    for ch in up:
        if ch in _GR2LAT:
            out.append(_GR2LAT[ch])
        elif ch.isascii() and ch.isalnum():
            out.append(ch)
        else:
            out.append(" ")
    words = "".join(out).split()
    slug = "_".join(words[:max_words])[:max_len].strip("_")
    return slug


def get_api_key():
    """Επιστρέφει το API key από env var GEMI_API_KEY, αλλιώς το default."""
    return os.environ.get("GEMI_API_KEY", DEFAULT_API_KEY).strip()


def session(api_key=None):
    s = requests.Session()
    s.headers.update({"api_key": api_key or get_api_key(), "Accept": "application/json"})
    return s


def list_activities(api_key=None, current_only=True):
    """
    Επιστρέφει τη λίστα ΚΑΔ (metadata/activities).
    current_only=True: μόνο η τρέχουσα ταξινομία (kad_2026), ώστε να μην
    εμφανίζονται παλιοί κωδικοί που δεν δουλεύουν πια.
    """
    s = session(api_key)
    r = s.get(f"{BASE}/metadata/activities", timeout=60)
    r.raise_for_status()
    data = r.json()
    if current_only:
        cur = [a for a in data if a.get("kadVersion") == KAD_VERSION]
        if cur:  # ασφάλεια: αν αλλάξει το ταξινομικό όνομα, μη μείνει κενό
            data = cur
    return sorted(data, key=lambda x: str(x.get("id", "")))


def search_activities(keyword, api_key=None, current_only=True):
    """
    Επιστρέφει τους ΚΑΔ των οποίων η περιγραφή περιέχει τη λέξη-κλειδί,
    αγνοώντας τόνους και πεζά/κεφαλαία. Κενή λέξη -> όλος ο κατάλογος.
    """
    kw = strip_acc(keyword)
    acts = list_activities(api_key, current_only=current_only)
    hits = [a for a in acts if kw in strip_acc(a.get("descr", ""))]
    return hits


def resolve_prefectures(prefectures):
    """
    Μετατρέπει την επιλογή νομών σε συμβολοσειρά κωδικών για το API.

    prefectures: "ALL" / None / κενή λίστα -> None (όλη η Ελλάδα)
                 λίστα με ονόματα ή κωδικούς -> "19,5,..."
    """
    if prefectures == "ALL" or not prefectures:
        return None
    ids = []
    for p in prefectures:
        ids.append(PREF_MAP.get(str(p).upper().strip(), str(p).strip()))
    return ",".join(ids)


def is_primary(company, prefix):
    for a in (company.get("activities") or []):
        act = a.get("activity") or {}
        if a.get("type") == "Κύρια" and str(act.get("id", "")).startswith(prefix):
            return True
    return False


def _excluded(name, exclude):
    up = (name or "").upper()
    return any(t in up for t in exclude)


def fetch_kad(kad, prefectures="ALL", api_key=None, primary_only=True,
              exclude=None, progress=None, should_stop=None, on_progress=None):
    """
    Αντλεί όλες τις εγγραφές για έναν ΚΑΔ.

    kad          : 8ψήφιος ΚΑΔ (string)
    prefectures  : "ALL" / λίστα ονομάτων ή κωδικών νομών
    api_key      : προαιρετικό override
    primary_only : κράτα μόνο εγγραφές με τον ΚΑΔ ως Κύρια δραστηριότητα
    exclude      : λίστα αποκλεισμού brand (default: DEFAULT_EXCLUDE)
    progress     : προαιρετικό callable(message:str) για κείμενο προόδου
    should_stop  : προαιρετικό callable()->bool για ακύρωση από τον χρήστη
    on_progress  : προαιρετικό callable(scanned:int, total:int) για μπάρα προόδου

    Επιστρέφει λίστα από γραμμές (κάθε γραμμή με τη σειρά του HDR).
    """
    if exclude is None:
        exclude = DEFAULT_EXCLUDE
    pref_ids = resolve_prefectures(prefectures)
    prefix = kad[:4]
    rows, off, total, seen = [], 0, None, set()
    s = session(api_key)

    def log(msg):
        if progress:
            progress(msg)

    while total is None or off < total:
        if should_stop and should_stop():
            log("  Ακυρώθηκε από τον χρήστη.")
            break
        url = (f"{BASE}/companies?activities={kad}&isActive=true"
               f"&resultsSize={PAGE}&resultsOffset={off}")
        if pref_ids:
            url += f"&prefectures={pref_ids}"
        try:
            r = s.get(url, timeout=60)
        except Exception as e:
            log(f"  σφάλμα δικτύου, επανάληψη: {e}")
            time.sleep(8)
            continue
        if r.status_code == 429:
            log("  429 (όριο ρυθμού), αναμονή...")
            time.sleep(9)
            continue
        if r.status_code == 404:
            log(f"  404 — ο ΚΑΔ {kad} δεν βρέθηκε (ίσως άλλαξε η ταξινομία).")
            break
        if r.status_code != 200:
            log(f"  status {r.status_code}, επανάληψη")
            time.sleep(9)
            continue
        j = r.json()
        total = j.get("searchMetadata", {}).get("totalCount", 0)
        arr = j.get("searchResults", []) or []
        for c in arr:
            if primary_only and not is_primary(c, prefix):
                continue
            name = (c.get("coNameEl") or "").strip()
            if _excluded(name, exclude):
                continue
            afm = c.get("afm") or ""
            if afm and afm in seen:
                continue
            if afm:
                seen.add(afm)
            pr = c.get("prefecture") or {}
            pr = pr.get("descr", "") if isinstance(pr, dict) else (pr or "")
            addr = ((c.get("street") or "") + " " + (c.get("streetNumber") or "")).strip()
            rows.append([
                name, afm, addr, c.get("city") or "", c.get("zipCode") or "",
                pr, c.get("phone") or "", c.get("email") or "", c.get("url") or "",
            ])
        off += len(arr)
        log(f"  {kad}: κρατήθηκαν {len(rows)} / σαρώθηκαν {off}/{total}")
        if on_progress:
            on_progress(off, total or off)
        if len(arr) < PAGE:
            break
        time.sleep(SLEEP)
    if on_progress:
        on_progress(total or off, total or off)  # 100% στο τέλος
    return rows


def _sanitize(name, used):
    x = (name or "ΛΟΙΠΑ").translate({ord(ch): "-" for ch in ':\\/?*[]'})[:28] or "ΛΟΙΠΑ"
    b, i = x, 1
    while x in used:
        x = f"{b}_{i}"
        i += 1
    used.add(x)
    return x


def summary(rows):
    """Επιστρέφει λίστα (Νομός, Πλήθος) φθίνουσα, χρήσιμη για UI."""
    groups = {}
    for r in rows:
        n = (r[5] or "ΛΟΙΠΑ").strip() or "ΛΟΙΠΑ"
        groups[n] = groups.get(n, 0) + 1
    return sorted(groups.items(), key=lambda kv: -kv[1])


def build_xlsx(rows, path):
    """Δημιουργεί το .xlsx: φύλλο «Σύνοψη» + ένα tab ανά Νομό."""
    groups = {}
    for r in rows:
        n = (r[5] or "ΛΟΙΠΑ").strip() or "ΛΟΙΠΑ"
        groups.setdefault(n, []).append(r)
    names = sorted(groups, key=lambda n: -len(groups[n]))
    wb = Workbook()
    ws = wb.active
    ws.title = "Σύνοψη"
    ws.append(["Νομός", "Πλήθος"])
    for n in names:
        ws.append([n, len(groups[n])])
    ws.append(["ΣΥΝΟΛΟ", len(rows)])
    hf = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor=NAVY)
    for c in ws[1]:
        c.font = hf
        c.fill = fill
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 10
    used = set()
    widths = [44, 11, 30, 16, 7, 16, 12, 26, 20]
    for n in names:
        sh = wb.create_sheet(_sanitize(n, used))
        sh.append(HDR)
        for c in sh[1]:
            c.font = hf
            c.fill = fill
            c.alignment = Alignment(horizontal="center")
        for row in groups[n]:
            sh.append(row)
        for i, w in enumerate(widths, 1):
            sh.column_dimensions[get_column_letter(i)].width = w
        sh.freeze_panes = "A2"
    wb.save(path)


def export_kad(kad, prefectures="ALL", out_dir="output", api_key=None,
               primary_only=True, exclude=None, progress=None, should_stop=None,
               on_progress=None, label=None):
    """
    Αντλεί έναν ΚΑΔ και γράφει το αντίστοιχο .xlsx.
    label: περιγραφή ΚΑΔ (ελληνικά) — μπαίνει στο όνομα αρχείου ως λατινικό slug.
    Επιστρέφει (path, rows).
    """
    os.makedirs(out_dir, exist_ok=True)
    rows = fetch_kad(kad, prefectures=prefectures, api_key=api_key,
                     primary_only=primary_only, exclude=exclude,
                     progress=progress, should_stop=should_stop,
                     on_progress=on_progress)
    slug = latin_slug(label) if label else ""
    name = f"GEMH_{kad}_{slug}_ana_Nomo.xlsx" if slug else f"GEMH_{kad}_ana_Nomo.xlsx"
    path = os.path.join(out_dir, name)
    build_xlsx(rows, path)
    return path, rows
